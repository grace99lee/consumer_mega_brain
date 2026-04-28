from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from exporters.base import BaseExporter
from models.schemas import AnalysisResult, Review, SentimentLabel


class ExcelExporter(BaseExporter):
    HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def export(
        self,
        result: AnalysisResult,
        reviews: list[Review],
        output_dir: Path,
    ) -> list[Path]:
        output_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = output_dir / "report.xlsx"

        wb = Workbook()

        # Summary sheet (rename default)
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, result)

        # Reviews sheet
        ws_reviews = wb.create_sheet("Reviews")
        self._write_reviews(ws_reviews, reviews)

        # Themes sheet
        ws_themes = wb.create_sheet("Themes")
        self._write_themes(ws_themes, result)

        # Quotes sheet
        ws_quotes = wb.create_sheet("Key Quotes")
        self._write_quotes(ws_quotes, result)

        # Unmet Needs sheet
        ws_needs = wb.create_sheet("Unmet Needs")
        self._write_needs(ws_needs, result)

        # Personas sheet
        ws_personas = wb.create_sheet("Personas")
        self._write_personas(ws_personas, result)

        # Sentiment sheet
        ws_sentiment = wb.create_sheet("Sentiment")
        self._write_sentiment(ws_sentiment, result)

        wb.save(str(xlsx_path))
        return [xlsx_path]

    def _style_header(self, ws, headers: list[str], row: int = 1):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

    def _auto_width(self, ws, max_width: int = 50):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    def _write_summary(self, ws, result: AnalysisResult):
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Consumer Insights Report: {result.query}"
        ws["A1"].font = Font(bold=True, size=16)

        info = [
            ("Query Type", result.query_type),
            ("Total Reviews", result.total_reviews),
            ("Sources", ", ".join(s.value for s in result.sources_used)),
            ("Overall Sentiment", f"{result.sentiment.overall.value} ({result.sentiment.score:+.2f})"),
            ("Generated", result.generated_at.strftime("%Y-%m-%d %H:%M")),
        ]

        for i, (label, value) in enumerate(info, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=str(value))

        ws.cell(row=len(info) + 4, column=1, value="Executive Summary").font = Font(bold=True, size=12)
        ws.merge_cells(f"A{len(info) + 5}:D{len(info) + 5}")
        summary_cell = ws.cell(row=len(info) + 5, column=1, value=result.executive_summary)
        summary_cell.alignment = Alignment(wrap_text=True)

        self._auto_width(ws)

    def _write_reviews(self, ws, reviews: list[Review]):
        headers = ["ID", "Source", "Author", "Text", "Rating", "Date", "URL", "Product"]
        self._style_header(ws, headers)

        for i, r in enumerate(reviews, 2):
            ws.cell(row=i, column=1, value=r.id)
            ws.cell(row=i, column=2, value=r.source.value)
            ws.cell(row=i, column=3, value=r.author or "")
            text_cell = ws.cell(row=i, column=4, value=r.text[:500])
            text_cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=5, value=r.rating)
            ws.cell(row=i, column=6, value=r.date.isoformat() if r.date else "")
            ws.cell(row=i, column=7, value=r.url or "")
            ws.cell(row=i, column=8, value=r.product_name or "")

        self._auto_width(ws)

    def _write_themes(self, ws, result: AnalysisResult):
        headers = ["Theme", "Description", "Mentions", "Positive", "Negative", "Neutral", "Mixed"]
        self._style_header(ws, headers)

        for i, t in enumerate(result.themes, 2):
            ws.cell(row=i, column=1, value=t.name)
            ws.cell(row=i, column=2, value=t.description).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=3, value=t.review_count)
            ws.cell(row=i, column=4, value=t.sentiment_breakdown.get(SentimentLabel.POSITIVE, t.sentiment_breakdown.get("positive", 0)))
            ws.cell(row=i, column=5, value=t.sentiment_breakdown.get(SentimentLabel.NEGATIVE, t.sentiment_breakdown.get("negative", 0)))
            ws.cell(row=i, column=6, value=t.sentiment_breakdown.get(SentimentLabel.NEUTRAL, t.sentiment_breakdown.get("neutral", 0)))
            ws.cell(row=i, column=7, value=t.sentiment_breakdown.get(SentimentLabel.MIXED, t.sentiment_breakdown.get("mixed", 0)))

        self._auto_width(ws)

    def _write_quotes(self, ws, result: AnalysisResult):
        headers = ["Quote", "Source", "Author", "Theme", "Sentiment"]
        self._style_header(ws, headers)

        for i, q in enumerate(result.key_quotes, 2):
            ws.cell(row=i, column=1, value=q.quote).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=2, value=q.source.value)
            ws.cell(row=i, column=3, value=q.author or "Anonymous")
            ws.cell(row=i, column=4, value=q.theme)
            ws.cell(row=i, column=5, value=q.sentiment.value)

        self._auto_width(ws)

    def _write_needs(self, ws, result: AnalysisResult):
        headers = ["Unmet Need", "Frequency", "Opportunity Score"]
        self._style_header(ws, headers)

        for i, n in enumerate(result.unmet_needs, 2):
            ws.cell(row=i, column=1, value=n.need).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=2, value=n.frequency)
            ws.cell(row=i, column=3, value=round(n.opportunity_score, 2))

        self._auto_width(ws)

    def _write_personas(self, ws, result: AnalysisResult):
        headers = ["Name", "Description", "Demographics", "Motivations", "Pain Points", "Prevalence"]
        self._style_header(ws, headers)

        for i, p in enumerate(result.personas, 2):
            ws.cell(row=i, column=1, value=p.name)
            ws.cell(row=i, column=2, value=p.description).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=3, value=p.demographics_hints)
            ws.cell(row=i, column=4, value="; ".join(p.motivations)).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=5, value="; ".join(p.pain_points)).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=6, value=p.estimated_prevalence)

        self._auto_width(ws)

    def _write_sentiment(self, ws, result: AnalysisResult):
        s = result.sentiment

        ws.cell(row=1, column=1, value="Sentiment Overview").font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="Overall").font = Font(bold=True)
        ws.cell(row=3, column=2, value=s.overall.value)
        ws.cell(row=4, column=1, value="Score").font = Font(bold=True)
        ws.cell(row=4, column=2, value=round(s.score, 2))

        # Distribution
        ws.cell(row=6, column=1, value="Distribution").font = Font(bold=True, size=12)
        headers = ["Sentiment", "Count"]
        self._style_header(ws, headers, row=7)
        row = 8
        for label, count in s.distribution.items():
            label_str = label.value if isinstance(label, SentimentLabel) else str(label)
            ws.cell(row=row, column=1, value=label_str)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # By source
        row += 1
        ws.cell(row=row, column=1, value="By Source").font = Font(bold=True, size=12)
        row += 1
        headers2 = ["Source", "Score"]
        self._style_header(ws, headers2, row=row)
        row += 1
        for source, score in s.by_source.items():
            source_str = source.value if hasattr(source, 'value') else str(source)
            ws.cell(row=row, column=1, value=source_str)
            ws.cell(row=row, column=2, value=round(score, 2))
            row += 1

        self._auto_width(ws)
